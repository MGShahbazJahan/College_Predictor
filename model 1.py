from tkinter import *
import openpyxl
from tkinter import messagebox

class predictor:
    def my_accuracy(self):
        racc=Tk()
        racc.title("Accuracy of Result")
        self.lal=Label(racc,text=str((self.c/self.t)*100))
        self.lal.pack()
        return None
    def about(self):
        with open("eaccuracy.txt","r") as f:
            ac=f.readlines()
        oacc=Tk()
        oacc.title("about the predictor")
        self.lac1=Label(oacc,text="over all accuracy is"+ac[0])
        self.lac1.pack()
        self.lac2=Label(oacc,text="number of results"+ac[1])
        self.lac2.pack()
        return None
    def ecalculate(self):
        self.ltotal=[]
        self.dtotal={}
        g=self.r.get()
        i=self.course.get()
        r=self.er.get()
        try:
            r=int(r)
        except ValueError:
            messagebox.showerror("Error","Invalid rank")
        #2018 calculations done frm here
        if(r>0):
            flag=0
            c=self.category.get()
            wb=openpyxl.load_workbook("eamcet2018.xlsx")
            s=wb['Table 1']
            if(g==1 and c=='OC'):
                m='J'
            if(g==2 and c=='OC'):
                m='I'
            if(g==1 and c=='BC-A'):
                m='L'
            if(g==2 and c=='BC-A'):
                m='K'
            if(g==1 and c=='BC-B'):
                m='N'
            if(g==2 and c=='BC-B'):
                m='M'
            if(g==1 and c=='BC-C'):
                m='P'
            if(g==2 and c=='BC-C'):
                m='O'
            if(g==1 and c=='BC-D'):
                m='R'
            if(g==2 and c=='BC-D'):
                m='Q'
            if(g==1 and c=='BC-E'):
                m='T'
            if(g==2 and c=='BC-E'):
                m='S'
            if(g==1 and c=='SC'):
                m='V'
            if(g==2 and c=='SC'):
                m='U'
            if(g==1 and c=='ST'):
                m='X'
            if(g==2 and c=='ST'):
                m='W'
            d2018={}
            l2018=[]
            for a in range(3,833):
                j=m+str(a)
                co=s[j].value
                if (co=='NA'):
                    continue
                j="B"+str(a)
                n=s[j].value
                j='G'+str(a)
                p=s[j].value
                if(p==i and r<co):
                    d2018[co]=n
                    l2018.append(co)
                    flag=1
            l2018.sort()
        #2019 calculations done frm here
        if(r>0):
            flag=0
            c=self.category.get()
            wb=openpyxl.load_workbook("eamcet2019.xlsx")
            s=wb['Table 1']
            if(g==1 and c=='OC'):
                m='J'
            if(g==2 and c=='OC'):
                m='I'
            if(g==1 and c=='BC-A'):
                m='L'
            if(g==2 and c=='BC-A'):
                m='K'
            if(g==1 and c=='BC-B'):
                m='N'
            if(g==2 and c=='BC-B'):
                m='M'
            if(g==1 and c=='BC-C'):
                m='P'
            if(g==2 and c=='BC-C'):
                m='O'
            if(g==1 and c=='BC-D'):
                m='R'
            if(g==2 and c=='BC-D'):
                m='Q'
            if(g==1 and c=='BC-E'):
                m='T'
            if(g==2 and c=='BC-E'):
                m='S'
            if(g==1 and c=='SC'):
                m='V'
            if(g==2 and c=='SC'):
                m='U'
            if(g==1 and c=='ST'):
                m='X'
            if(g==2 and c=='ST'):
                m='W'
            d2019={}
            l2019=[]
            for a in range(3,833):
                j=m+str(a)
                co=s[j].value
                if (co=='NA'):
                    continue
                j="B"+str(a)
                n=s[j].value
                j='G'+str(a)
                p=s[j].value
                if(p==i and r<co):
                    d2019[co]=n
                    l2019.append(co)
                    flag=1
            l2019.sort()
        #2017 calculations done frm here
        if(r>0):
            flag=0
            c=self.category.get()
            wb=openpyxl.load_workbook("eamcet2017.xlsx")
            s=wb['Table 1']
            if(g==1 and c=='OC'):
                m='J'
            if(g==2 and c=='OC'):
                m='I'
            if(g==1 and c=='BC-A'):
                m='L'
            if(g==2 and c=='BC-A'):
                m='K'
            if(g==1 and c=='BC-B'):
                m='N'
            if(g==2 and c=='BC-B'):
                m='M'
            if(g==1 and c=='BC-C'):
                m='P'
            if(g==2 and c=='BC-C'):
                m='O'
            if(g==1 and c=='BC-D'):
                m='R'
            if(g==2 and c=='BC-D'):
                m='Q'
            if(g==1 and c=='BC-E'):
                m='T'
            if(g==2 and c=='BC-E'):
                m='S'
            if(g==1 and c=='SC'):
                m='V'
            if(g==2 and c=='SC'):
                m='U'
            if(g==1 and c=='ST'):
                m='X'
            if(g==2 and c=='ST'):
                m='W'
            d2017={}
            l2017=[]
            for a in range(3,833):
                j=m+str(a)
                co=s[j].value
                if (co=='NA'):
                    continue
                j="B"+str(a)
                n=s[j].value
                j='G'+str(a)
                p=s[j].value
                if(p==i and r<co):
                    d2017[co]=n
                    l2017.append(co)
                    flag=1
            l2017.sort()
        #2016 calculations done frm here
        if(r>0):
            flag=0
            c=self.category.get()
            wb=openpyxl.load_workbook("eamcet2016.xlsx")
            s=wb['Table 1']
            if(g==1 and c=='OC'):
                m='J'
            if(g==2 and c=='OC'):
                m='I'
            if(g==1 and c=='BC-A'):
                m='L'
            if(g==2 and c=='BC-A'):
                m='K'
            if(g==1 and c=='BC-B'):
                m='N'
            if(g==2 and c=='BC-B'):
                m='M'
            if(g==1 and c=='BC-C'):
                m='P'
            if(g==2 and c=='BC-C'):
                m='O'
            if(g==1 and c=='BC-D'):
                m='R'
            if(g==2 and c=='BC-D'):
                m='Q'
            if(g==1 and c=='BC-E'):
                m='T'
            if(g==2 and c=='BC-E'):
                m='S'
            if(g==1 and c=='SC'):
                m='V'
            if(g==2 and c=='SC'):
                m='U'
            if(g==1 and c=='ST'):
                m='X'
            if(g==2 and c=='ST'):
                m='W'
            d2016={}
            l2016=[]
            for a in range(3,833):
                j=m+str(a)
                co=s[j].value
                if (co=='NA'):
                    continue
                j="B"+str(a)
                n=s[j].value
                j='G'+str(a)
                p=s[j].value
                if(p==i and r<co):
                    d2016[co]=n
                    l2016.append(co)
                    flag=1
            l2016.sort()
            for x in range(10):
                self.ltotal.append(d2018[l2018[x]])
            for x in range(10):
                self.ltotal.append(d2019[l2019[x]])
            for x in range(10):
                self.ltotal.append(d2017[l2017[x]])
            for x in range(10):
                self.ltotal.append(d2016[l2016[x]])
            self.c=0
            self.t=0
            for x in range(10):
                self.c=self.c+self.ltotal.count(d2019[l2019[x]])
                self.t=self.t+4
            for x in range(10):
                self.c=self.c+self.ltotal.count(d2018[l2018[x]])
                self.t=self.t+4
            for x in range(10):
                self.c=self.c+self.ltotal.count(d2017[l2017[x]])
                self.t=self.t+4
            for x in range(10):
                self.c=self.c+self.ltotal.count(d2016[l2016[x]])
                self.t=self.t+4
            with open("eaccuracy.txt","r") as f:
                ac=f.readlines()
            ac[0]=float(ac[0])
            ac[1]=float(ac[1])
            ac[0]=((self.c/self.t)*100+(ac[1]*ac[0]))/(ac[1]+1)
            ac[1]=ac[1]+1
            with open("eaccuracy.txt","w") as f:
                f.write(str(ac[0])+'\n')
                f.write(str(ac[1]))  
            if(flag==0):
                messagebox.showerror("Error","You could not crack EAMCET\nBetter luck next time")
            else:
                rclg=Tk()
                rclg.title("List of colleges")
                rclg.geometry("750x600")
                if(len(l2018)>=1):
                    c0=Button(rclg,text=d2018[l2018[0]],padx=300,pady=15)
                    c0.pack()
                    if(len(l2018)>=2):
                        c1=Button(rclg,text=d2018[l2018[1]],padx=300,pady=15)
                        c1.pack()
                        if(len(l2018)>=3):
                            c2=Button(rclg,text=d2018[l2018[2]],padx=300,pady=15)
                            c2.pack()
                            if(len(l2018)>=4):
                                c3=Button(rclg,text=d2018[l2018[3]],padx=300,pady=15)
                                c3.pack()
                                if(len(l2018)>=5):
                                    c4=Button(rclg,text=d2018[l2018[4]],padx=300,pady=15)
                                    c4.pack()
                                    if(len(l2018)>=6):
                                        c5=Button(rclg,text=d2018[l2018[5]],padx=300,pady=15)
                                        c5.pack()
                                        if(len(l2018)>=7):
                                            c6=Button(rclg,text=d2018[l2018[6]],padx=300,pady=15)
                                            c6.pack()
                                            if(len(l2018)>=8):
                                                c7=Button(rclg,text=d2018[l2018[7]],padx=300,pady=15)
                                                c7.pack()
                                                if(len(l2018)>=9):
                                                    c8=Button(rclg,text=d2018[l2018[8]],padx=300,pady=15)
                                                    c8.pack()
                                                    if(len(l2018)>=10):
                                                        c9=Button(rclg,text=d2018[l2018[9]],padx=300,pady=15)
                                                        c9.pack()
                button_acc=Button(rclg,text="click here to check accuracy of the result",command=self.my_accuracy)
                button_acc.pack()
        else:
            messagebox.showerror("Error","Invalid rank")
        return None
    
    def jcalculate(self):
        g=self.r.get()
        g1="Female-only (including Supernumerary)"
        g2="Gender-Neutral"
        i=self.course.get()
        flag=0
        r=self.er.get()
        try:
            r=int(r)
        except ValueError:
            messagebox.showerror("Error","Invalid rank")
        if(r>0):
            c=self.category.get()
            wb=openpyxl.load_workbook("jee.xlsx")
            s=wb['Sheet1']
            l=[]
            d={}
            for a in range(2,9339):
                j='G'+str(a)
                co=s[j].value
                j='B'+str(a)
                p=s[j].value
                j='D'+str(a)
                q=s[j].value
                j='E'+str(a)
                u=s[j].value
                r=int(r)
                if (type(co)!=int):
                    continue
                if(co>r and p==i and c==q and g==1):
                    if(u==g2 or u==g1):
                        k='A'+str(a)
                        n=s[k].value
                        d[co]=n
                        l.append(co)
                        flag=1
                else:
                    if(co>r and p==i and c==q and g==2):
                        if(u==g2):
                            k='A'+str(a)
                            n=s[k].value
                            d[co]=n
                            l.append(co)
                            flag=1
            l.sort()
            if(flag==0):
                messagebox.showerror("Error","You could not crack EAMCET\nBetter luck next time")
            rclg=Tk()
            rclg.title("List of colleges")
            rclg.geometry("750x540")
            if(len(l)>=1):
                    c0=Button(rclg,text=d[l[0]],padx=300,pady=15)
                    c0.pack()
                    if(len(l)>=2):
                        c1=Button(rclg,text=d[l[1]],padx=300,pady=15)
                        c1.pack()
                        if(len(l)>=3):
                            c2=Button(rclg,text=d[l[2]],padx=300,pady=15)
                            c2.pack()
                            if(len(l)>=4):
                                c3=Button(rclg,text=d[l[3]],padx=300,pady=15)
                                c3.pack()
                                if(len(l)>=5):
                                    c4=Button(rclg,text=d[l[4]],padx=300,pady=15)
                                    c4.pack()
                                    if(len(l)>=6):
                                        c5=Button(rclg,text=d[l[5]],padx=300,pady=15)
                                        c5.pack()
                                        if(len(l)>=7):
                                            c6=Button(rclg,text=d[l[6]],padx=300,pady=15)
                                            c6.pack()
                                            if(len(l)>=8):
                                                c7=Button(rclg,text=d[l[7]],padx=300,pady=15)
                                                c7.pack()
                                                if(len(l)>=9):
                                                    c8=Button(rclg,text=d[l[8]],padx=300,pady=15)
                                                    c8.pack()
                                                    if(len(l)>=10):
                                                        c9=Button(rclg,text=d[l[9]],padx=300,pady=15)
                                                        c9.pack()
        else:
            messagebox.showerror("Error","Invalid rank")
        return None
        
    def jee(self):
        self.root.destroy()
        self.root=Tk()
        self.root.title("JEE Predictions")
        self.l1=Label(self.root,text="Giving your exam rank will help us recommend you better colleges. If you don't have actual rank, then enter expected rank.")
        self.l2=Label(self.root,text="        ")
        self.l3=Label(self.root,text="         ")
        self.l4=Label(self.root,text="        ")
        self.l5=Label(self.root,text="Enter your gender")
        self.l6=Label(self.root,text="Are you specially abled")
        self.l7=Label(self.root,text="            ")
        self.l8=Label(self.root,text="            ")
        self.l9=Label(self.root,text="        ")
        self.l10=Label(self.root,text="        ")
        self.l11=Label(self.root,text="        ")
        self.l12=Label(self.root,text="         ")
        self.button_e=Button(self.root,text="EAMCET",padx=160,pady=20,bg="white",font=("arial",15,"bold"),command= lambda:self.eamcet())
        self.button_j=Button(self.root,text="JEE MAINS",padx=160,pady=20,fg="white",bg="gray",font=("arial",15,"bold"),command= lambda:self.jee())
        self.button_e.grid(row=0,column=0,columnspan=2)
        self.button_j.grid(row=0,column=2,columnspan=2)
        self.er=Entry(self.root,width=40,borderwidth=20)
        self.er.insert(0,"Enter your rank")
        self.l1.grid(row=1,column=0,columnspan=4)
        self.l11.grid(row=2,column=1,columnspan=2)
        self.category= StringVar()
        self.category.set("Enter your category here")
        self.di=OptionMenu(self.root,self.category,'OPEN','OBC-NCL','EWS','ST','SC')
        self.di.grid(row=5,column=1,columnspan=2)
        self.l2.grid(row=4,column=1,columnspan=2)
        self.er.grid(row=3,column=1,columnspan=2)
        self.l3.grid(row=6,column=1,columnspan=2)
        self.course= StringVar()
        self.course.set("Enter the course here")
        self.di=OptionMenu(self.root,self.course,'Computer Science and Engineering (4 Years, Bachelor of Technology)','Mechanical Engineering (4 Years, Bachelor of Technology)','Civil Engineering (4 Years, Bachelor of Technology)','Electronics and Electrical Communication Engineering (4 Years, Bachelor of Technology)','Electrical and Electronics Engineering (4 Years, Bachelor of Technology)','Artificial Intelligence (4 Years, Bachelor of Technology)','Data Science and Engineering (4 Years, Bachelor of Technology)','Information Technology (4 Years, Bachelor of Technology)','Chemical Engineering (4 Years, Bachelor of Technology)','Metallurgical Engineering (4 Years, Bachelor of Technology)','Mining Engineering (4 Years, Bachelor of Technology)','Aerospace Engineering (4 Years, Bachelor of Technology)','Bio Engineering (4 Years, Bachelor of Technology)',)
        self.di.grid(row=7,column=1,columnspan=2)
        self.l4.grid(row=8,column=1,columnspan=2)
        self.l4.grid(row=10,column=1,columnspan=2)
        self.r=IntVar()
        self.l5.grid(row=11,column=1,columnspan=2)
        Radiobutton(self.root,text="Female",variable=self.r,value=1).grid(row=12,column=1)
        Radiobutton(self.root,text="Male",variable=self.r,value=2).grid(row=12,column=2)
        self.r1=IntVar()
        self.l6.grid(row=13,column=1,columnspan=2)
        Radiobutton(self.root,text="Yes",variable=self.r1,value=1).grid(row=14,column=1)
        Radiobutton(self.root,text="No",variable=self.r1,value=2).grid(row=14,column=2)
        self.l10.grid(row=15,column=1,columnspan=2)
        self.l7.grid(row=16,column=1,columnspan=2)
        self.l9.grid(row=17,column=1,columnspan=2)
        self.l8.grid(row=18,column=1,columnspan=2)
        self.button_calc=Button(self.root,text="Predict Results",padx=90,pady=30,fg="white",bg="orange",font=("arial",15,"bold"),command=lambda:self.jcalculate())
        self.button_calc.grid(row=19,column=1,columnspan=2)
    def eamcet(self):
        self.root.destroy()
        self.root=Tk()
        self.root.title("EAMCET Predictions")
        self.l1=Label(self.root,text="Giving your exam rank will help us recommend you better colleges. If you don't have actual rank, then enter expected rank.")
        self.l2=Label(self.root,text="        ")
        self.l3=Label(self.root,text="         ")
        self.l4=Label(self.root,text="        ")
        self.l5=Label(self.root,text="Enter your gender")
        self.l6=Label(self.root,text="Are you specially abled")
        self.l7=Label(self.root,text="            ")
        self.l8=Label(self.root,text="            ")
        self.l9=Label(self.root,text="        ")
        self.l10=Label(self.root,text="        ")
        self.l11=Label(self.root,text="        ")
        self.l12=Label(self.root,text="         ")
        self.button_e=Button(self.root,text="EAMCET",padx=160,pady=20,fg="white",bg="gray",font=("arial",15,"bold"),command= lambda:self.eamcet())
        self.button_j=Button(self.root,text="JEE MAINS",padx=160,pady=20,bg="white",font=("arial",15,"bold"),command= lambda:self.jee())
        self.button_e.grid(row=0,column=0,columnspan=2)
        self.button_j.grid(row=0,column=2,columnspan=2)
        self.er=Entry(self.root,width=40,borderwidth=20)
        self.er.insert(0,"Enter your rank (ex-50000)")
        self.button_abtacc=Button(self.root,text="about",command= lambda: self.about())
        self.button_abtacc.grid(row=1,column=3,columnspan=2)
        self.l1.grid(row=2,column=0,columnspan=4)
        self.l11.grid(row=3,column=1,columnspan=2)
        self.l2.grid(row=4,column=1,columnspan=2)
        self.er.grid(row=5,column=1,columnspan=2)
        self.l3.grid(row=6,column=1,columnspan=2)
        self.category = StringVar()
        self.category.set("Enter your category here")
        self.di=OptionMenu(self.root,self.category,'OC','BC-A','BC-B','BC-C','BC-D','BC-E','SC','ST')
        self.di.grid(row=7,column=1,columnspan=2)
        self.l12.grid(row=8,column=1,columnspan=2)
        self.course= StringVar()
        self.course.set("Enter the course here")
        self.di=OptionMenu(self.root,self.course,'CSE', 'MEC', 'ECE', 'EEE', 'MET','CIV', 'CHE','TEX', 'MIN','AGR', 'BIO')
        self.di.grid(row=9,column=1,columnspan=2)
        self.l4.grid(row=10,column=1,columnspan=2)
        self.r=IntVar()
        self.l5.grid(row=11,column=1,columnspan=2)
        Radiobutton(self.root,text="Female",variable=self.r,value=1).grid(row=12,column=1)
        Radiobutton(self.root,text="Male",variable=self.r,value=2).grid(row=12,column=2)
        self.l10.grid(row=15,column=1,columnspan=2)
        self.l7.grid(row=16,column=1,columnspan=2)
        self.l9.grid(row=17,column=1,columnspan=2)
        self.l8.grid(row=18,column=1,columnspan=2)
        self.button_calc=Button(self.root,text="Predict Results",padx=90,pady=30,fg="white",bg="orange",font=("arial",15,"bold"),command=lambda:self.ecalculate())
        self.button_calc.grid(row=19,column=1,columnspan=2)
    def __init__(self):
        self.root=Tk()
        self.root.title("College Predictor")
        self.root.geometry("860x600")
        self.button_e=Button(self.root,text="EAMCET",padx=160,pady=20,bg="white",font=("arial",15,"bold"),command= lambda:self.eamcet())
        self.button_j=Button(self.root,text="JEE MAINS",padx=160,pady=20,bg="white",font=("arial",15,"bold"),command= lambda:self.jee())
        self.button_e.grid(row=0,column=0,columnspan=2)
        self.button_j.grid(row=0,column=2,columnspan=2)
        self.ld=Label(self.root,text="Select the examination",font=("arial",15,"bold"))
        self.ld.grid(row=3,column=0,columnspan=4)
obj=predictor()
obj.root.mainloop()
